"""Helpers for preparing SpreadsheetBench tasks inside SkillGen.

SpreadsheetBench is a real-world spreadsheet-manipulation benchmark (NeurIPS 2024,
https://github.com/RUCKBReasoning/SpreadsheetBench). Each instance contains:

- ``instruction``: a natural-language request about one or more sheets.
- ``spreadsheet_path``: a folder with OJ-style test cases. Each test case ships
  a ``{No.}_{id}_input.xlsx`` file together with a matching ``_answer.xlsx``.
- ``answer_position`` / ``answer_sheet``: the cell range / sheet that the model
  must populate.
- ``data_position``: the cell range holding the source data.

This adapter focuses on the lightweight integration used by SkillGen: it
downloads the 912-instance archive hosted on HuggingFace, extracts it on demand
and converts each item into the shared ``{instance_id, input, ground_truth,
metadata}`` schema. The heavy OJ evaluation (Docker + LibreOffice recalculation)
is NOT implemented here; it is documented in the upstream repo.
"""

from __future__ import annotations

import json
import tarfile
from pathlib import Path
from typing import Any

_REPO_ROOT = Path(__file__).resolve().parent
_DEFAULT_ROOT = _REPO_ROOT / "data" / "spreadsheetbench"

_HF_REPO = "KAKA22/SpreadsheetBench"
_ARCHIVE_FILES = {
    "full_912": "spreadsheetbench_912_v0.1.tar.gz",
    "verified_400": "spreadsheetbench_verified_400.tar.gz",
}
_EXTRACTED_DIRS = {
    "full_912": "all_data_912_v0.1",
    "verified_400": "spreadsheetbench_verified_400",
}
SPREADSHEETBENCH_VARIANTS = tuple(_ARCHIVE_FILES.keys())

_MAX_ROWS_PER_SHEET_DEFAULT = 20
_MAX_COLS_PER_SHEET_DEFAULT = 20


def _archive_path(variant: str, root_dir: Path) -> Path:
    return root_dir / _ARCHIVE_FILES[variant]


def _extracted_root(variant: str, root_dir: Path) -> Path:
    return root_dir / _EXTRACTED_DIRS[variant]


def _download_archive(variant: str, root_dir: Path) -> Path:
    """Download the tar.gz from HuggingFace if it is not already present."""
    archive = _archive_path(variant, root_dir)
    if archive.exists():
        return archive

    from huggingface_hub import hf_hub_download

    root_dir.mkdir(parents=True, exist_ok=True)
    downloaded = hf_hub_download(
        repo_id=_HF_REPO,
        filename=_ARCHIVE_FILES[variant],
        repo_type="dataset",
        local_dir=str(root_dir),
    )
    return Path(downloaded)


def _extract_archive(variant: str, root_dir: Path) -> Path:
    """Extract the dataset archive to ``root_dir`` (skips if already extracted)."""
    extracted = _extracted_root(variant, root_dir)
    dataset_json = extracted / "dataset.json"
    if dataset_json.exists():
        return extracted

    archive = _download_archive(variant, root_dir)
    with tarfile.open(archive, "r:gz") as tf:
        tf.extractall(path=root_dir)

    if not dataset_json.exists():
        candidates = list(root_dir.glob("*/dataset.json"))
        if not candidates:
            raise FileNotFoundError(
                f"Could not locate dataset.json after extracting {archive}."
            )
        extracted = candidates[0].parent
    return extracted


def ensure_spreadsheetbench_data(
    variant: str = "full_912",
    root_dir: str | Path | None = None,
) -> Path:
    """Download + extract the dataset and return the extracted folder."""
    if variant not in _ARCHIVE_FILES:
        supported = ", ".join(_ARCHIVE_FILES)
        raise ValueError(
            f"Unsupported SpreadsheetBench variant '{variant}'. Supported: {supported}"
        )
    root = Path(root_dir) if root_dir else _DEFAULT_ROOT
    return _extract_archive(variant, root)


def load_spreadsheetbench_records(
    variant: str = "full_912",
    root_dir: str | Path | None = None,
) -> list[dict[str, Any]]:
    """Return the raw list of SpreadsheetBench items from ``dataset.json``."""
    extracted = ensure_spreadsheetbench_data(variant, root_dir)
    with open(extracted / "dataset.json", encoding="utf-8") as f:
        return json.load(f)


def _format_cell(value: Any) -> str:
    if value is None:
        return ""
    try:
        import datetime as _dt
        if isinstance(value, _dt.datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S") if (value.hour or value.minute or value.second) else value.strftime("%Y-%m-%d")
        if isinstance(value, _dt.date):
            return value.strftime("%Y-%m-%d")
    except Exception:
        pass
    text = str(value)
    return text.replace("\n", " ").replace("\r", " ")


def _render_workbook_preview(
    xlsx_path: Path,
    max_rows: int = _MAX_ROWS_PER_SHEET_DEFAULT,
    max_cols: int = _MAX_COLS_PER_SHEET_DEFAULT,
) -> str:
    """Render a compact textual preview of an xlsx file.

    Returns an empty string if openpyxl is not available so the caller can
    fall back to an instruction-only prompt.
    """
    try:
        import openpyxl
    except ImportError:
        return ""

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as exc:
        return f"[failed to read {xlsx_path.name}: {exc}]"

    sections: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                dim = ws.calculate_dimension()
            except Exception:
                dim = f"A1:{getattr(ws, 'max_column', '?')}x{getattr(ws, 'max_row', '?')}"
            rows: list[list[str]] = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx >= max_rows:
                    rows.append([f"... ({row_idx}+ more rows)"])
                    break
                cells = [_format_cell(v) for v in list(row)[:max_cols]]
                if len(row) > max_cols:
                    cells.append(f"... (+{len(row) - max_cols} cols)")
                rows.append(cells)
            table = "\n".join("\t".join(cells) for cells in rows) if rows else "(empty sheet)"
            sections.append(f"### Sheet: {sheet_name} (range={dim})\n{table}")
    finally:
        wb.close()
    return "\n\n".join(sections)


def _discover_test_cases(instance_dir: Path) -> tuple[list[str], list[str]]:
    input_files: list[str] = []
    answer_files: list[str] = []
    if not instance_dir.exists():
        return input_files, answer_files
    for path in sorted(instance_dir.iterdir()):
        name = path.name
        if name.endswith("_input.xlsx"):
            input_files.append(name)
        elif name.endswith("_answer.xlsx"):
            answer_files.append(name)
    return input_files, answer_files


def build_spreadsheetbench_prompt(
    item: dict[str, Any],
    data_dir: Path,
    max_rows: int = _MAX_ROWS_PER_SHEET_DEFAULT,
    max_cols: int = _MAX_COLS_PER_SHEET_DEFAULT,
) -> str:
    instance_dir = data_dir / item["spreadsheet_path"]
    input_files, _ = _discover_test_cases(instance_dir)

    parts = [
        "You are given a real-world spreadsheet manipulation task.",
        f"Task ID: {item.get('id')}",
        f"Instruction type: {item.get('instruction_type')}",
        f"Target sheet: {item.get('answer_sheet')}",
        f"Cells to populate (answer_position): {item.get('answer_position')}",
        f"Source data range (data_position): {item.get('data_position')}",
        "",
        "Instruction:",
        item.get("instruction", "").strip(),
    ]

    if input_files:
        preview_file = instance_dir / input_files[0]
        preview = _render_workbook_preview(preview_file, max_rows=max_rows, max_cols=max_cols)
        if preview:
            parts.extend([
                "",
                f"Preview of the first test-case input workbook ({preview_file.name}):",
                preview,
            ])
        parts.extend([
            "",
            f"All {len(input_files)} test-case input files live at: {instance_dir}",
        ])

    parts.extend([
        "",
        "You MUST produce a Python 3 solution with the following EXACT contract:",
        "",
        "```python",
        "def solve(input_path: str, output_path: str) -> None:",
        "    \"\"\"Read the input workbook at `input_path`, apply the manipulation",
        "    requested above, and write the resulting workbook to `output_path`.",
        "    The produced workbook must contain the target sheet with the requested",
        "    cell range populated with the correct values.\"\"\"",
        "    # ... your implementation here ...",
        "```",
        "",
        "Requirements:",
        "- Use `openpyxl` to read and write .xlsx files.",
        "- Do NOT hard-code any filenames; always use the `input_path` / `output_path` arguments.",
        "- Do NOT call `solve()` yourself; the grader will invoke it per test case.",
        "- Preserve the sheet structure; only change cells within the requested answer range.",
        "- Write ACTUAL VALUES (not formulas) into the answer cells so the grader can",
        "  read them back with `data_only=True`. Only emit formulas if the statement",
        "  explicitly requires them.",
        "",
        "Return a single fenced Python code block containing the `solve` function.",
    ])
    return "\n".join(parts)


def convert_spreadsheetbench_item(
    item: dict[str, Any],
    idx: int,
    data_dir: Path | None = None,
    max_rows: int = _MAX_ROWS_PER_SHEET_DEFAULT,
    max_cols: int = _MAX_COLS_PER_SHEET_DEFAULT,
) -> dict[str, Any]:
    data_dir = data_dir or _DEFAULT_ROOT / _EXTRACTED_DIRS["full_912"]
    instance_dir = data_dir / item["spreadsheet_path"]
    input_files, answer_files = _discover_test_cases(instance_dir)
    prompt = build_spreadsheetbench_prompt(item, data_dir, max_rows=max_rows, max_cols=max_cols)

    return {
        "instance_id": str(item.get("id") or idx),
        "input": prompt,
        "ground_truth": None,
        "metadata": {
            "benchmark": "spreadsheetbench",
            "instruction_type": item.get("instruction_type"),
            "answer_position": item.get("answer_position"),
            "answer_sheet": item.get("answer_sheet"),
            "data_position": item.get("data_position"),
            "spreadsheet_path": item.get("spreadsheet_path"),
            "spreadsheet_abs_dir": str(instance_dir),
            "test_case_count": len(input_files),
            "input_files": input_files,
            "answer_files": answer_files,
        },
    }
