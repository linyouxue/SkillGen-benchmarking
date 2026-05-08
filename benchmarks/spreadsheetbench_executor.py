"""SpreadsheetBench executor + grader.

Given a model's natural-language answer for a SpreadsheetBench task, this module:
  1. Extracts a Python solution from the response.
  2. Standardises it to expose `solve(input_path: str, output_path: str) -> None`.
  3. For each of the N test cases attached to the task, runs the solution in a
     subprocess with a wall-clock timeout (default 30s) against the test-case
     input workbook, then compares the produced workbook against the official
     answer workbook over the cell range given by `metadata["answer_position"]`.

A test case passes iff every cell in the expected answer range matches the
corresponding cell in the produced workbook (after type-aware normalisation).
The final score is `passed_test_cases / total_test_cases`; `passed` is True iff
all test cases pass.

This mirrors the shape of `livecodebench_adapter.evaluate_livecodebench_output`
so `trajectory.evaluate_trajectory` can dispatch on `metadata["benchmark"]`.
"""

from __future__ import annotations

import io
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
import textwrap
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


_DEFAULT_TIMEOUT_S = 30
_DEFAULT_TOL = 1e-6
_MAX_CELLS_REPORTED = 10  # cap on mismatched cells in error_message


# Code extraction + standardisation

_FENCE_RE = re.compile(
    r"```(?:python|py)?\s*\n(.*?)```",
    re.DOTALL | re.IGNORECASE,
)


def _extract_python_code(raw: str) -> str:
    """Pull the largest Python code block out of a model response.

    Falls back to treating the whole response as code if no fenced block is
    present. Always returns a non-empty string (possibly empty if the model
    produced nothing). Never raises.
    """
    if not raw:
        return ""
    blocks = _FENCE_RE.findall(raw)
    if blocks:
        # Pick the longest one (models sometimes emit multiple, e.g. explanation
        # snippet + full solution).
        return max(blocks, key=len).strip()
    # No fence: trust the raw body, but strip obvious chatty prefixes/suffixes.
    return raw.strip()


_SOLVE_DEF_RE = re.compile(
    # Matches `def solve(input_path, ...)` OR `def solve(input_path: str, ...)`
    # at column 0 only (module level). Tolerates type annotations and default values.
    r"^def\s+solve\s*\(\s*input_path\b[^)]*\)",
    re.MULTILINE,
)


def _ensure_solve_entrypoint(code: str) -> str:
    """If the model's code already defines `solve(input_path, ...)` at module
    scope, return it unchanged. Otherwise wrap the snippet into a module-level
    `solve(input_path, output_path)` function that executes the snippet with
    those two names bound.

    The primary expected path is that the model follows the prompt contract
    (it does ~99% of the time), so the fallback exists only to survive the
    rare cases where the model emits a plain script.
    """
    if _SOLVE_DEF_RE.search(code):
        return code

    indented = textwrap.indent(code.rstrip(), "    ")
    return (
        "def solve(input_path, output_path):\n"
        "    __name__ = '__main__'\n"
        f"{indented}\n"
    )


# Subprocess driver

_RUNNER_TEMPLATE = r"""
import json, sys, traceback, os
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")
os.environ.setdefault("OMP_NUM_THREADS", "1")

# BEGIN MODEL CODE
{MODEL_CODE}
# END MODEL CODE

def _run():
    input_path  = {INPUT_PATH!r}
    output_path = {OUTPUT_PATH!r}
    try:
        solve(input_path, output_path)
    except Exception as exc:
        return {{"ok": False, "error": f"{{type(exc).__name__}}: {{exc}}",
                 "trace": traceback.format_exc()}}
    if not os.path.exists(output_path):
        return {{"ok": False, "error": "solve() did not produce output_path",
                 "trace": ""}}
    return {{"ok": True}}

if __name__ == "__main__":
    result = _run()
    print(json.dumps(result))
"""


def _run_solution_once(
    code: str,
    input_path: Path,
    output_path: Path,
    *,
    timeout_s: int,
) -> dict[str, Any]:
    """Execute `code` in a subprocess. Return {ok: bool, error: str|None, trace: str}."""
    runner_src = _RUNNER_TEMPLATE.format(
        MODEL_CODE=code,
        INPUT_PATH=str(input_path),
        OUTPUT_PATH=str(output_path),
    )

    with tempfile.TemporaryDirectory(prefix="sb_run_") as tmpdir:
        runner_file = Path(tmpdir) / "runner.py"
        runner_file.write_text(runner_src)
        try:
            proc = subprocess.run(
                [sys.executable, str(runner_file)],
                capture_output=True,
                text=True,
                timeout=timeout_s,
                check=False,
            )
        except subprocess.TimeoutExpired:
            return {"ok": False, "error": f"timeout after {timeout_s}s", "trace": ""}
        except Exception as exc:  # pragma: no cover - defensive
            return {"ok": False, "error": f"subprocess failed: {exc}", "trace": ""}

    stdout = (proc.stdout or "").strip()
    if proc.returncode != 0 and not stdout:
        stderr = (proc.stderr or "").strip()[:2000]
        return {"ok": False,
                "error": f"subprocess exit {proc.returncode}",
                "trace": stderr}

    # Take the last JSON-looking line from stdout (model code may also print).
    json_line = None
    for line in reversed(stdout.splitlines()):
        line = line.strip()
        if line.startswith("{") and line.endswith("}"):
            json_line = line
            break
    if json_line is None:
        return {"ok": False,
                "error": "runner produced no JSON",
                "trace": (proc.stderr or "")[:2000]}
    try:
        return json.loads(json_line)
    except json.JSONDecodeError:
        return {"ok": False, "error": "malformed runner JSON", "trace": json_line[:500]}


# Cell comparison

def _normalise_cell(v: Any, *, tol: float = _DEFAULT_TOL):
    """Normalise a cell value for tolerant comparison.

    - None / "" / blank -> None
    - numeric-like -> float rounded to tolerance
    - string -> stripped, lowercased
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return bool(v)
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:  # NaN
            return "nan"
        return round(float(v), 6)
    s = str(v).strip()
    if s == "":
        return None
    # Try numeric parsing
    try:
        f = float(s.replace(",", ""))
        return round(f, 6)
    except ValueError:
        pass
    return s.lower()


def _compare_cell_range(
    produced_path: Path,
    expected_path: Path,
    cell_range: str | None,
    sheet_name: str | None,
) -> dict[str, Any]:
    """Compare the `cell_range` region of `produced_path` against `expected_path`.

    Returns {"match": bool, "mismatches": [(addr, produced, expected), ...],
             "error": str|None}.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries
    except ImportError as exc:
        return {"match": False, "mismatches": [], "error": f"openpyxl missing: {exc}"}

    try:
        wb_prod = load_workbook(produced_path, data_only=True, read_only=True)
    except Exception as exc:
        return {"match": False, "mismatches": [],
                "error": f"cannot open produced workbook: {type(exc).__name__}: {exc}"}
    try:
        wb_exp = load_workbook(expected_path, data_only=True, read_only=True)
    except Exception as exc:
        wb_prod.close()
        return {"match": False, "mismatches": [],
                "error": f"cannot open expected workbook: {type(exc).__name__}: {exc}"}

    try:
        def _pick_sheet(wb, name):
            if name and name in wb.sheetnames:
                return wb[name]
            return wb.active

        ws_prod = _pick_sheet(wb_prod, sheet_name)
        ws_exp = _pick_sheet(wb_exp, sheet_name)

        if cell_range:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        else:
            # Compare the full used range of the expected sheet.
            min_col, min_row = 1, 1
            max_col = ws_exp.max_column or 1
            max_row = ws_exp.max_row or 1

        mismatches: list[tuple[str, Any, Any]] = []
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                exp_val = _normalise_cell(ws_exp.cell(r, c).value)
                prod_val = _normalise_cell(ws_prod.cell(r, c).value)
                if exp_val != prod_val:
                    from openpyxl.utils import get_column_letter
                    addr = f"{get_column_letter(c)}{r}"
                    mismatches.append((addr, prod_val, exp_val))
                    if len(mismatches) > _MAX_CELLS_REPORTED:
                        break
            if len(mismatches) > _MAX_CELLS_REPORTED:
                break

        return {"match": not mismatches,
                "mismatches": mismatches,
                "error": None}
    finally:
        wb_prod.close()
        wb_exp.close()


# Public API

def evaluate_spreadsheetbench_output(
    model_output: str,
    instance_metadata: dict[str, Any],
    *,
    timeout_s: int = _DEFAULT_TIMEOUT_S,
) -> dict[str, Any]:
    """Grade a single SpreadsheetBench instance.

    Returns a dict with at least:
      - passed: bool              -- True iff ALL test cases matched
      - score: float              -- passed_tc / total_tc
      - total_test_cases: int
      - passed_test_cases: int
      - error_message: str|None
      - extracted_code: str
      - per_test_case: list[{case_id, ok, error, mismatches_head}]
    """
    extracted = _extract_python_code(str(model_output or ""))
    prepared = _ensure_solve_entrypoint(extracted)

    meta = instance_metadata or {}
    abs_dir = meta.get("spreadsheet_abs_dir")
    input_files = meta.get("input_files") or []
    answer_files = meta.get("answer_files") or []
    cell_range = meta.get("answer_position")
    sheet_name = meta.get("answer_sheet")

    if not abs_dir or not input_files or not answer_files:
        return {
            "passed": False,
            "score": 0.0,
            "total_test_cases": 0,
            "passed_test_cases": 0,
            "error_message": "missing spreadsheet_abs_dir / input_files / answer_files",
            "extracted_code": extracted,
            "per_test_case": [],
        }

    abs_dir_p = Path(abs_dir)
    pairs: list[tuple[Path, Path, str]] = []
    # Pair each input with its corresponding answer by numeric prefix (1_, 2_, 3_).
    def _key(fname: str) -> str:
        return fname.split("_", 1)[0]
    inp_by_key = {_key(f): f for f in input_files}
    ans_by_key = {_key(f): f for f in answer_files}
    for key in sorted(set(inp_by_key) & set(ans_by_key), key=lambda x: int(x) if x.isdigit() else x):
        pairs.append((abs_dir_p / inp_by_key[key], abs_dir_p / ans_by_key[key], key))

    if not pairs:
        return {
            "passed": False,
            "score": 0.0,
            "total_test_cases": 0,
            "passed_test_cases": 0,
            "error_message": "no aligned input/answer test-case pairs",
            "extracted_code": extracted,
            "per_test_case": [],
        }

    per_tc: list[dict[str, Any]] = []
    passed_count = 0
    first_error: str | None = None

    for input_path, answer_path, case_id in pairs:
        with tempfile.TemporaryDirectory(prefix="sb_case_") as tmpdir:
            tmp_input = Path(tmpdir) / input_path.name
            tmp_output = Path(tmpdir) / f"{case_id}_produced.xlsx"
            shutil.copyfile(input_path, tmp_input)

            run_res = _run_solution_once(prepared, tmp_input, tmp_output,
                                         timeout_s=timeout_s)
            if not run_res.get("ok"):
                per_tc.append({
                    "case_id": case_id,
                    "ok": False,
                    "error": run_res.get("error") or "run failed",
                    "trace": (run_res.get("trace") or "")[:500],
                    "mismatches_head": [],
                })
                if first_error is None:
                    first_error = f"case {case_id}: {run_res.get('error')}"
                continue

            cmp_res = _compare_cell_range(tmp_output, answer_path,
                                          cell_range=cell_range,
                                          sheet_name=sheet_name)
            if cmp_res.get("error"):
                per_tc.append({
                    "case_id": case_id,
                    "ok": False,
                    "error": cmp_res["error"],
                    "trace": "",
                    "mismatches_head": [],
                })
                if first_error is None:
                    first_error = f"case {case_id}: {cmp_res['error']}"
                continue

            ok = bool(cmp_res["match"])
            if ok:
                passed_count += 1
            per_tc.append({
                "case_id": case_id,
                "ok": ok,
                "error": None if ok else "cell mismatch",
                "trace": "",
                "mismatches_head": [
                    {"cell": a, "produced": p, "expected": e}
                    for (a, p, e) in (cmp_res.get("mismatches") or [])[:_MAX_CELLS_REPORTED]
                ],
            })
            if not ok and first_error is None:
                head = cmp_res.get("mismatches") or []
                if head:
                    a, p, e = head[0]
                    first_error = f"case {case_id}: cell {a} produced={p!r} expected={e!r}"
                else:
                    first_error = f"case {case_id}: cell mismatch"

    total = len(pairs)
    passed_all = passed_count == total and total > 0
    return {
        "passed": bool(passed_all),
        "score": float(passed_count) / float(total) if total else 0.0,
        "total_test_cases": total,
        "passed_test_cases": passed_count,
        "error_message": None if passed_all else (first_error or "test cases failed"),
        "extracted_code": extracted,
        "per_test_case": per_tc,
    }


__all__ = ["evaluate_spreadsheetbench_output"]
